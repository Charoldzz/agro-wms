from collections import Counter
from pathlib import Path
import re

import openpyxl

SOURCE = Path(r"C:\Users\Harold\Desktop\tagribol\AlmacenEjemplo2.xlsx")

PACKAGE_RE = re.compile(
    r"(?:x\s*)?(\d+(?:[\.,]\d+)?)\s*(lts?|lt\.?|litros?|kgs?|kg\.?|grs?|gr\.?|ml\.?|cc)\b",
    re.IGNORECASE,
)


def main():
    workbook = openpyxl.load_workbook(SOURCE, data_only=True)
    sheet = workbook["Almacen"]
    headers = [cell.value for cell in sheet[1]]
    index = {header: position for position, header in enumerate(headers)}

    parsed = []
    missing = []
    stock_rows = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = str(row[index["nomb_cata"]] or "")
        stock = float(row[index["saldf_fin"]] or 0)
        if stock > 0:
            stock_rows += 1

        matches = list(PACKAGE_RE.finditer(name))
        if matches:
            match = matches[-1]
            parsed.append((name, match.group(1), match.group(2), stock))
        else:
            missing.append(name)

    print("rows", sheet.max_row - 1)
    print("stock_gt_0", stock_rows)
    print("parsed_package", len(parsed))
    print("missing_package", len(missing))
    print("units", Counter(unit.lower().replace(".", "") for _, _, unit, _ in parsed).most_common())
    print("sizes", Counter((size, unit.lower().replace(".", "")) for _, size, unit, _ in parsed).most_common(25))
    print("\nsample parsed")
    for item in parsed[:25]:
        print(item)
    print("\nsample missing")
    for item in missing[:25]:
        print(item)


if __name__ == "__main__":
    main()
